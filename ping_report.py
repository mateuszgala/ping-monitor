from ping3 import ping
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from datetime import datetime


with open("hosts.txt", "r", encoding="utf-8") as file:
    hosts = [line.strip() for line in file if line.strip()]


wb = Workbook()
ws = wb.active
ws.title = "Ping Report"
ws.append(["Host", "Status", "Date", "Time"])


green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")



for host in hosts:
    try:
        response = ping(host, timeout=1)
        status = "OK" if response else "OFFLINE"
    except Exception:
        status = "PING ERROR"

    date = datetime.now().strftime("%Y-%m-%d")
    time = datetime.now().strftime("%H:%M:%S")
    ws.append([host, status, date, time])

    fill = green if status == "OK" else red
    for cell in ws[ws.max_row]:
        cell.fill = fill


wb.save("report.xlsx")
print("✅ Report saved as report.xlsx")
